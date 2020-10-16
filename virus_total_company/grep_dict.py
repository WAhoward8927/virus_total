"""取得API回傳結果並客製結果"""
import os
import time
import all_func
import globals
# from virus_total import all_func, globals
import openpyxl
from openpyxl import load_workbook


def grep_data(web_response, path, spare=None):
    """結果過濾並寫檔"""
    response_to_json = web_response.json()
    fp = open(path, 'w+')
    try:
        analysis_data = response_to_json['data']['attributes']['last_analysis_results']
        status = response_to_json['data']['attributes']['last_analysis_stats']
        count = int()  # 各狀態加總
        malicious_num = status.get('malicious')  # 有害數量
        suspicious_num = status.get('suspicious')  # 可疑數量
        for status_name, status_value in status.items():
            count += status_value
        combined_calc = malicious_num + suspicious_num  # 給加入、解管用的值
        fp.writelines(f"({malicious_num}/{count})\n")
        fp.writelines(f"可疑數量：{suspicious_num}\n")
        call_count = 0
        if bool(analysis_data) is False and call_count == 0:  # 只讓程式再呼叫一次
            for repeat in range(1):
                time.sleep(15)
                all_func.General().urls(spare)  # 如果last_analysis_results是空的就在重跑一次(針對url)
                call_count += 1
        else:
            pass
        top = response_to_json['data']
        scan_type = response_to_json['data']['type']  # 掃描類型
        if scan_type == 'url':
            source_id = response_to_json['data']['attributes']['url']  # 掃描url的結果排版不同,需使用此方法
        else:
            source_id = response_to_json['data']['id']  # 本次掃描資料
        del response_to_json['data']['attributes']
        title = list()
        index = list()
        index.append(source_id), index.append(scan_type)
        for id_key in top:
            title.append(id_key)
        title.remove('links')
        new_dict = dict(zip(title, index))
        for sort_key, sort_value in new_dict.items():  # 寫入id與type
            fp.writelines(f"{sort_key}: {sort_value}\n")
        category = list()
        company_name = list()
        for analysis_key, analysis_value in analysis_data.items():
            category.append(analysis_value.get('category'))
            company_name.append(analysis_key)
            detail_category = analysis_value.get('category')  # 取出危害分類
            detail_method = analysis_value.get('method')  # 取出掃描方法
            detail_result = analysis_value.get('result')  # 取出掃描結果
            fp.writelines(f"{analysis_key} -> {detail_category}, {detail_method}, {detail_result}\n")
        merge_to_dict = dict(zip(company_name, category))
        combine_result = list()  # 有檢測出問題的廠商
        for name, threat_level in merge_to_dict.items():
            '''if 'malicious' == threat_level:'''
            if threat_level != 'harmless' and threat_level != 'undetected':  # 暫時改為非"無害"或"未檢測"皆寫入廠商名
                combine_result.append(name)
        combine_result_to_str = ', '.join(combine_result)
        target_mode = int(globals.get_demo_value())  # 從globals get取得值,然後轉int給下方使用
        
        if target_mode == 1:  # 模式選擇區塊
            _write_excel('./加入黑名單.xlsx', scan_type, source_id, combined_calc, count,
                         combine_result_to_str, target_mode)  # mode1的掃描總數暫時改為"有害"+"可疑"
            '''_write_excel('./加入黑名單.xlsx', scan_type, source_id, malicious_num, count,
                         combine_result_to_str, target_mode)'''
        elif target_mode == 2:
            _write_excel('./解管黑名單.xlsx', scan_type, source_id, combined_calc, count,
                         combine_result_to_str, target_mode)
        else:
            return print("無此功能")

    except KeyError:
        fp.writelines(web_response.text)
        

def _write_excel(path, sheet_name, target, quantity, count, harmful, mode):
    """帶入：路徑檔名、工作表名、廠商名、危害總數、掃出有危害的廠商"""
    if os.path.isfile(path):  # 如果檔案存在
        wb = load_workbook(path)  # 載入此檔
        all_sheetname = wb.sheetnames  # 抓出所有工作表名(data type:list)
        if sheet_name in all_sheetname:  # 如果全部表名內有目前指定的表名
            _write_data(path, wb, sheet_name, target, quantity, count, harmful, mode)
        else:
            wb.create_sheet(sheet_name)
            _write_data(path, wb, sheet_name, target, quantity, count, harmful, mode)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        _write_data(path, wb, sheet_name, target, quantity, count, harmful, mode)


def _write_data(path, wb, sheet_name, target, quantity, count, harmful, mode):
    """將結果寫入"""
    """傳入：檔案路徑、workbook功能、表名、掃描目標、危害數量、掃到危害的廠商"""
    def _fixed():
        designation_sheet['A1'] = '掃描目標'
        designation_sheet['B1'] = '危害數量'
        designation_sheet['C1'] = '總廠商數'
        designation_sheet['D1'] = '掃到危害的廠商'
    if mode == 1:  # 模式1
        designation_sheet = wb[sheet_name]  # 指定工作表
        _fixed()
        if quantity != 0:  # 只要"有害"或"可疑">0就寫excel
            designation_sheet.append([target, quantity, count, harmful])
            wb.save(path)
        else:
            pass
    else:  # 模式2
        designation_sheet = wb[sheet_name]
        _fixed()
        if quantity == 0:  # "有害"或"可疑"需恆等於0就會寫excel
            designation_sheet.append([target, quantity, count])
            wb.save(path)
        else:  # 否則寫此格式或不寫
            pass
