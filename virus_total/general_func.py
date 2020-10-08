"""依檔名與副檔名呼叫指定副程式"""
import os
import time
import csv
import xlrd
from openpyxl import load_workbook
import all_func
import catch_file
# from virus_total import all_func, catch_file


def file_check():
    """分派副檔名,支援csv、txt、xlsx、xls"""
    check_result = catch_file.file_check()
    if len(check_result) == 0:
        return print("沒找到規範中的檔名或附檔名的檔案")
    for judgment_extension in check_result:
        print("\r程式運行中", end='')
        extension = os.path.splitext(judgment_extension)[-1]  # 檢查副檔名，呼叫不同的讀取文件方式
        if extension in '.csv':
            excel_csv(judgment_extension)
        elif extension in '.txt':
            read_txt(judgment_extension)
        elif extension in 'xlsx':
            excel_xlsx(judgment_extension)
        else:
            excel_xls(judgment_extension)


def judgment_func(file_name, data):
    """分派檔名"""
    if 'domaindata' in file_name:
        all_func.General().domain(data)
    elif 'urldata' in file_name:
        all_func.General().urls(data)
    elif 'filedata' in file_name:
        all_func.General().file(data)
    elif 'ipdata' in file_name:
        all_func.General().ip_scan(data)
    else:
        return print(f"程式不接受{file_name}的檔案名稱")


def excel_csv(file):
    """讀取csv方法"""
    open_file = open(file, 'r', newline='', encoding='utf-8-sig')
    rows = csv.reader(open_file, delimiter=',')
    for top_shelf in rows:  # 從csv逐行讀出是list
        for second_shelf in top_shelf:  # 需再一次loop輸出str
            judgment_func(file, second_shelf)
            time.sleep(15)  # 免費帳號的等待時間


def excel_xls(file):
    """讀取xls方法(可讀取多分頁)"""
    wb = xlrd.open_workbook(file)  # 讀取指定xls檔
    sheets = wb.sheets()  # 掃出該檔所有sheet
    for sheet in sheets:
        rows = sheet.get_rows()  # 取得所有列
        for row in rows:
            judgment_func(file, row[0].value)
            time.sleep(15)  # 免費帳號的等待時間


def excel_xlsx(file):
    """讀取xlsx方法(可讀取多分頁)"""
    wb = load_workbook(file)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        rows = ws.rows
        for row in rows:  # 出來是tuple,需要再次迭代
            for col in row:
                judgment_func(file, col.value)
                time.sleep(15)


def read_txt(file):
    """讀取txt的方法，一列一筆資料後直接換行"""
    openfile = open(file, 'r')
    lines = openfile.readlines()
    for i in range(len(lines)):
        line = lines[i].replace("\n", "")
        judgment_func(file, line)
        time.sleep(15)
